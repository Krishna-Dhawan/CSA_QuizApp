"""Krishna Dhawan, 2023A7PS0111H"""
import openpyxl as opx
import tkinter as tk
import tkinter.font
from tkinter import messagebox
import random

qs = opx.load_workbook("Questions.xlsx")
sheet = qs['Sheet1']
sheet.cell(row=2, column=10).value = 0      # Using a cell to store the score
root = tk.Tk()


def Verify():
    def log():
        user = name_entry.get()
        passwor = pass_entry.get()

        if user == "admin" and passwor == "123":
            add_Ques()
            newWin.destroy()
        else:
            messagebox.showwarning("Invalid details", "Incorrect user or password")

    newWin = tk.Toplevel(root)
    newWin.title("Login")

    name_label = tk.Label(newWin, text="User: ", font=tk.font.Font(size=36))
    name_label.grid(row=0, column=0)
    name_entry = tk.Entry(newWin, width=65)
    name_entry.grid(row=0, column=1)

    pass_label = tk.Label(newWin, text="Password: ", font=tk.font.Font(size=36))
    pass_label.grid(row=1, column=0)
    pass_entry = tk.Entry(newWin, width=65)
    pass_entry.grid(row=1, column=1)

    login = tk.Button(newWin, text="Submit", command=log)
    login.grid(row=2, column=0)


def add_Ques():
    def Sub():
        q = q_entry.get()
        o1 = o1e.get()
        o2 = o2e.get()
        o3 = o3e.get()
        o4 = o4e.get()
        ans = int(g.get())
        addQues(q, [o1, o2, o3, o4], ans)
        qWin.destroy()

    qWin = tk.Toplevel(root)
    qWin.title("new question")

    q_label = tk.Label(qWin, text="Prompt: ", font=tk.font.Font(size=36))
    q_label.grid(row=0, column=0)
    q_entry = tk.Entry(qWin, width=65)
    q_entry.grid(row=0, column=1)

    o1l = tk.Label(qWin, text="Option 1", font=tk.font.Font(size=32))
    o1l.grid(row=1, column=0)
    o1e = tk.Entry(qWin, width=65)
    o1e.grid(row=1, column=1)

    o2l = tk.Label(qWin, text="Option 2", font=tk.font.Font(size=32))
    o2l.grid(row=2, column=0)
    o2e = tk.Entry(qWin, width=65)
    o2e.grid(row=2, column=1)

    o3l = tk.Label(qWin, text="Option 3", font=tk.font.Font(size=32))
    o3l.grid(row=3, column=0)
    o3e = tk.Entry(qWin, width=65)
    o3e.grid(row=3, column=1)

    o4l = tk.Label(qWin, text="Option 4", font=tk.font.Font(size=32))
    o4l.grid(row=4, column=0)
    o4e = tk.Entry(qWin, width=65)
    o4e.grid(row=4, column=1)

    g = tk.StringVar()
    tk.Radiobutton(qWin, text="Option 1", variable=g, value="1").grid(row=5, column=0)
    tk.Radiobutton(qWin, text="Option 2", variable=g, value="2").grid(row=5, column=1)
    tk.Radiobutton(qWin, text="Option 3", variable=g, value="3").grid(row=5, column=2)
    tk.Radiobutton(qWin, text="Option 4", variable=g, value="4").grid(row=5, column=3)

    submit_new = tk.Button(qWin, text="Submit", command=Sub)
    submit_new.grid(row=6, column=0)


def addQues(ques, opt, correct):
    a = 1
    while 1:
        if sheet.cell(row=a, column=1).value is not None:
            a += 1
        else:
            break
    sheet.cell(row=a, column=1).value = a
    sheet.cell(row=a, column=2).value = ques
    for b in range(4):
        sheet.cell(row=a, column=b+3).value = opt[b]
    sheet.cell(row=a, column=7).value = correct

    sheet.cell(row=1, column=10).value += 1


def Begin():
    def Chek():
        val = True
        try:
            num = int(b_entry.get())
        except ValueError:
            messagebox.showerror("Invalid Entry", "Enter an integer")
            val = False
        if val:
            if num <= sheet.cell(row=1, column=10).value:
                chooseQues(num)
                bWin.destroy()
            else:
                messagebox.showwarning("Invalid Entry",
                                       "Enter a number less than " + str(sheet.cell(row=1, column=10).value))
                val = False

    bWin = tk.Toplevel(root)
    bWin.title("Begin")

    b_label = tk.Label(bWin, text="Number of Questions: ", font=tk.font.Font(size=36))
    b_label.grid(row=0, column=0)
    b_entry = tk.Entry(bWin, width=65)
    b_entry.grid(row=0, column=1)

    start = tk.Button(bWin, text="Submit", command=Chek)
    start.grid(row=1, column=0)


def chooseQues(num):
    lis = []
    for i in range(int(sheet.cell(row=1, column=10).value)):
        lis.append(i+1)
    for i in range(num):
        x = random.choice(lis)
        displayQues(i+1, x, num)
        lis.remove(x)


def displayQues(n, x, num):
    def Next():
        if int(g.get()) == int(sheet.cell(row=x, column=7).value):
            sheet.cell(row=2, column=10).value += 1
        if n == num:
            disScore(num)
        disWin.destroy()

    disWin = tk.Toplevel(root)
    disWin.title("Ques " + str(n))

    qtxt = str(sheet.cell(row=x, column=2).value)
    ql = tk.Label(disWin, text=qtxt, font=tk.font.Font(size=36))
    ql.grid(row=0, column=0)

    g = tk.StringVar()
    tk.Radiobutton(disWin, text=str(sheet.cell(row=x, column=3).value), variable=g, value="1").grid(row=1, column=0)
    tk.Radiobutton(disWin, text=str(sheet.cell(row=x, column=4).value), variable=g, value="2").grid(row=1, column=1)
    tk.Radiobutton(disWin, text=str(sheet.cell(row=x, column=5).value), variable=g, value="3").grid(row=2, column=0)
    tk.Radiobutton(disWin, text=str(sheet.cell(row=x, column=6).value), variable=g, value="4").grid(row=2, column=1)

    nxt = tk.Button(disWin, text="Next", command=Next,
                    height=5, width=20)
    nxt.grid(row=3, column=0)


def disScore(num):
    sWin = tk.Toplevel(root)
    sWin.title("Result")

    lab = tk.Label(sWin, text=str(sheet.cell(row=2, column=10).value) + "/" + str(num), font=tk.font.Font(size=36))
    lab.grid(row=0, column=0)


# Main Window
QuesMaker = tk.Button(root, text="QuesMaker", command=Verify,
                      height=5, width=20)
QuesMaker.grid(row=0, column=0)
Student = tk.Button(root, text="Student", command=Begin,
                    height=5, width=20)
Student.grid(row=0, column=1)

root.mainloop()

qs.save("Questions.xlsx")
